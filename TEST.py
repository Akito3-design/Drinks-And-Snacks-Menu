# from tkinter import * 
# from tkinter.ttk import Combobox 
 
# window = Tk() 
# window.title("Процент совместимости") 
# window.geometry('500x400') 
 
# # Первая строка выбора 
# combo1 = Combobox(window, width=20) # Увеличили ширину 
# combo1['values'] = (1, 2, 3, 4, 5, "Текст") 
# combo1.current(1) 
# combo1.grid(column=0, row=0, padx=10, pady=10) # Добавили отступы 
 
# # Кнопка "Сравнить!" 
# def clicked(): 
#   value1 = combo1.get() 
#   value2 = combo2.get() 
 
#   # Очищаем текст в окне вывода 
#   output_label.config(text="") 
 
#   # Добавляем текст в окно вывода 
#   output_label.config(text=f"Вы выбрали в первой строке: {value1}\nВы выбрали во второй строке: {value2}") 
 
# btn = Button(window, text="Сравнить!", command=clicked, width=10) # Увеличили ширину 
# btn.grid(column=1, row=1, padx=10, pady=10) # Добавили отступы 
 
# # Вторая строка выбора 
# combo2 = Combobox(window, width=20) # Увеличили ширину 
# combo2['values'] = (1, 2, 3, 4, 5, "Другой текст") 
# combo2.current(0) 
# combo2.grid(column=0, row=2, padx=10, pady=10) # Добавили отступы 
 
# # Окно вывода 
# output_label = Label(window, text="", wraplength=300) # Добавлен wraplength для переноса строки 
# output_label.grid(column=0, row=3, columnspan=2, padx=10, pady=10) # Размещаем на двух колонках 
 
# window.mainloop()

import openpyxl

def просмотреть_файл():
    """Просматривает содержимое файла Excel и позволяет выбрать лист, а также возвращаться к выбору листа."""
    
    файл = input("Введите Напитки или Закуски: ")
    try:
        wb = openpyxl.load_workbook(файл + ".xlsx")  # Добавляем расширение .xlsx

        while True:
            print("\nДоступный выбор:")
            for index, sheet in enumerate(wb.sheetnames):
                print(f"{index + 1}. {sheet}")
            print("0. Назад")

            лист_выбор = int(input("\nВведите номер листа для просмотра или 0 для возврата: ")) - 1

            if лист_выбор == -1:
                break  # Возвращаемся в главное меню
            elif 0 <= лист_выбор < len(wb.sheetnames):
                sheet = wb.worksheets[лист_выбор]  # Выбираем лист по индексу
                print("\nМеню:", sheet.title)
                for row in sheet.iter_rows(values_only=True):
                    print(row)
            else:
                print("Неверный номер листа.")
                
    except FileNotFoundError:
        print("Файл не найден.")

def главное_меню():
    """Отображает главное меню."""
    
    print("\nГлавное меню:")
    print("1. Посмотреть меню")
    print("2. Выход")
    
    выбор = input("Выберите пункт меню: ")
    return выбор

if __name__ == "__main__":
    while True:
        выбор = главное_меню()
        if выбор == '1':
            просмотреть_файл()
        elif выбор == '2':
            break
        else:
            print("Неверный выбор.")


