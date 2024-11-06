import openpyxl

def просмотреть_файл():
    """Просматривает содержимое файла Excel и позволяет выбрать лист, а также возвращаться к выбору листа."""
    
    файл = input("Введите имя файла Excel: ")
    try:
        wb = openpyxl.load_workbook(файл + ".xlsx")  # Добавляем расширение .xlsx

        while True:
            print("\nДоступные листы:")
            for index, sheet in enumerate(wb.sheetnames):
                print(f"{index + 1}. {sheet}")
            print("0. Назад")

            лист_выбор = int(input("\nВведите номер листа для просмотра или 0 для возврата: ")) - 1

            if лист_выбор == -1:
                break  # Возвращаемся в главное меню
            elif 0 <= лист_выбор < len(wb.sheetnames):
                sheet = wb.worksheets[лист_выбор]  # Выбираем лист по индексу
                print("\nЛист:", sheet.title)
                for row in sheet.iter_rows(values_only=True):
                    print(row)
            else:
                print("Неверный номер листа.")
                
    except FileNotFoundError:
        print("Файл не найден.")

def главное_меню():
    """Отображает главное меню."""
    
    print("\nГлавное меню:")
    print("1. Просмотреть файл Excel")
    print("2. Выход")
    
    выбор = input("Выберите пункт меню: ")
    return выбор

if name == "main":
    while True:
        выбор = главное_меню()
        if выбор == '1':
            просмотреть_файл()
        elif выбор == '2':
            break
        else:
            print("Неверный выбор.")
